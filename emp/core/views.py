from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
import pandas as pd
from datetime import datetime, time
import re
from django.http import HttpResponse

# Default times (kept as datetime objects)
from collections import deque
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
# Set default times as time objects
# Default times (kept as datetime objects)
DEFAULT_TIMES = {
    "morning_check_in": datetime.strptime("09:00", "%H:%M"),

    "afternoon_check_out": datetime.strptime("20:00", "%H:%M")
}


def calculate_hours_with_error_correction(data):
    # Ensure 'Time' column is in datetime format
    data['Time'] = pd.to_datetime(data['Time'])
    data['Date'] = data['Time'].dt.date  # Add a Date column for grouping

    # Group by employee and date
    grouped = data.groupby(['id', 'Date'])

    results = []

    for (employee, date), group in grouped:
        group = group.sort_values('Time')  # Sort by time to analyze chronologically

        # Initialize variables
        morning_check_in, last_check_out = None, None
        unassigned = deque()
        
        # Initialize flags for this employee and date
        check_in_flag = False
        check_out_flag = False

        # Process records
        for _, row in group.iterrows():
            time = row['Time']
            status = row['In / Out Status']

            if status == "OT-In":
                if not morning_check_in:  # Take the first valid check-in as morning check-in
                    morning_check_in = time
                else:
                    unassigned.append((time, status))  # Add to unassigned if extra
                    print (list(unassigned))
            elif status == "OT-Out":
                # Update the last check-out to be the latest valid check-out time
                last_check_out = time

        # Apply defaults for missing times
        if morning_check_in:
            # Ensure morning check-in is before 11:00
            if morning_check_in.time() >= datetime.strptime("11:00", "%H:%M").time():
                morning_check_in = datetime.combine(date, DEFAULT_TIMES["morning_check_in"].time())
                check_in_flag = True  # Update flag
                print(f"Correction made for {employee} on {date}: Morning check-in adjusted to {morning_check_in.time()}.")
        else:
            # Default morning check-in if no valid check-in exists
            morning_check_in = datetime.combine(date, DEFAULT_TIMES["morning_check_in"].time())
            check_in_flag = True  # Update flag
            print(f"No valid morning check-in for {employee} on {date}. Default set to {morning_check_in.time()}.")

        if last_check_out:
            # Ensure last check-out is after 16:00
            if last_check_out.time() < datetime.strptime("15:00", "%H:%M").time():
                last_check_out = datetime.combine(date, DEFAULT_TIMES["afternoon_check_out"].time())
                check_out_flag = True  # Update flag
                print(f"Correction made for {employee} on {date}: Last check-out adjusted to {last_check_out.time()}.")
        else:
            # Default last check-out if no valid check-out exists
            last_check_out = datetime.combine(date, DEFAULT_TIMES["afternoon_check_out"].time())
            check_out_flag = True  # Update flag
            print(f"No valid last check-out for {employee} on {date}. Default set to {last_check_out.time()}.")


        # Log results for this employee and date
        results.append({
            "employee_id": employee,
            "date": date,
            "morning_check_in": morning_check_in.time(),
            "last_check_out": last_check_out.time(),
            "unassigned_actions": list(unassigned),  # Log unprocessed actions for debugging
            "check_in_flag": check_in_flag,
            "check_out_flag": check_out_flag,
        })

        # Print the flags for this employee and date
    print (list(unassigned))
    return results



@csrf_exempt
def edit_time(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_id = data.get('employee_id')
            date = data.get('date')
            new_time = data.get('time')
            time_type = data.get('type')

            # Parse the new time and update your dataset or database
            parsed_time = datetime.strptime(new_time, "%H:%M").time()

            # Update logic for database or in-memory data (example assumes in-memory)
            # You would typically update the database here
            # e.g., updating a database model instance
            print(f"Updating {time_type} for {employee_id} on {date} to {parsed_time}.")

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request method"})

def calculate_hours_with_tracking(data):
    results = calculate_hours_with_error_correction(data)  # Reuse correction function
    results_df = pd.DataFrame(results)

    # Add grouping columns
    results_df['day'] = results_df['date']
    results_df['week'] = results_df['date'].apply(lambda x: x.strftime('%Y-%U'))
    results_df['month'] = results_df['date'].apply(lambda x: x.strftime('%Y-%m'))

    return results_df


def aggregate_by_period(data):
    results_df = calculate_hours_with_tracking(data)

    # Group by day
    daily_data = results_df.groupby(['employee_id', 'day']).agg(
        morning_check_in=pd.NamedAgg(column="morning_check_in", aggfunc="first"),
        last_check_out=pd.NamedAgg(column="last_check_out", aggfunc="last"),
        date=pd.NamedAgg(column="day", aggfunc="first"),
        check_in_flag=pd.NamedAgg(column="check_in_flag", aggfunc="max"),  # Take max to determine if any were flagged
        check_out_flag=pd.NamedAgg(column="check_out_flag", aggfunc="max")   # Take max to determine if any were flagged
    ).reset_index().to_dict(orient="records")

    

    return {
        "daily": daily_data,
    }

def upload_file(request):
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        # Load and process the file
        df = pd.read_excel(file_path)
        df.rename(columns={
            'Jour': 'Date',
            'Heure': 'Time',
            'Status': 'In / Out Status',
            'Nom': 'id',
        }, inplace=True)

        # Convert 'Time' column to datetime format
        df['Time'] = pd.to_datetime(df['Time'])

        # Calculate results
        aggregated_data = aggregate_by_period(df)

        # Pass aggregated data to the template
        return render(request, 'results.html', {'aggregated_data': aggregated_data})

    return render(request, 'upload.html')









def convert_to_24hr_format(time_str):
    """Convert 12-hour formatted time to 24-hour formatted time."""
    time_str = time_str.lower().strip()  # Normalize the input time
    is_pm = False

    # Check if "p.m." or "a.m." exists
    if 'p.m.' in time_str or 'pm' in time_str:
        is_pm = True
        time_str = time_str.replace('p.m.', '').replace('pm', '').strip()
    elif 'a.m.' in time_str or 'am' in time_str:
        time_str = time_str.replace('a.m.', '').replace('am', '').strip()

    # If time only has hour part (e.g., 9), add ':00'
    if re.match(r"^\d+$", time_str):  # If only an hour is provided
        time_str = f"{time_str}:00"  # Add ":00" to make it "9:00"

    time_parts = time_str.split(":")
    if len(time_parts) != 2:
        raise ValueError("Invalid time format. Ensure the time is in HH:MM format or uses 'a.m./p.m.' notation.")

    hours, minutes = map(int, time_parts)

    # If PM and it's not 12:xx, add 12 to the hour
    if is_pm and hours != 12:
        hours += 12
    # If AM and it's 12:xx, convert hour to 0
    elif not is_pm and hours == 12:
        hours = 0

    return f"{hours:02}:{minutes:02}"

def calculate_hours_worked(row):
    """Calculate the total hours worked between check-in and check-out times."""
    check_in = datetime.strptime(row['morning_check_in'], "%H:%M")
    check_out = datetime.strptime(row['last_check_out'], "%H:%M")
    worked_hours = (check_out - check_in).seconds / 3600 - 1  # Subtract 1 hour for lunch or breaks
    return round(worked_hours, 2)

def assign_custom_month(date):
    """Assign a custom month based on a 26th-to-25th range."""
    if date.day >= 26:
        return date.month  # Current month
    else:
        return (date.month - 1) if date.month > 1 else 12  # Previous month (handle January to December case)

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def process_additional_file(request):
    """Process the uploaded file and calculate hours worked, weekly and monthly summaries."""
    if request.method == "POST" and request.FILES.get('additional_file'):
        file = request.FILES['additional_file']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        # Load the Excel file into a DataFrame
        df = pd.read_excel(file_path)

        # Rename columns for clarity
        df.rename(columns={
            'Employee ID': 'id',
            'Date': 'date',
            'Morning Check-In': 'morning_check_in',
            'Last Check-Out': 'last_check_out',
        }, inplace=True)

        # Convert the 'date' column to datetime format
        df['date'] = pd.to_datetime(df['date'])

        # Add the day name column
        df['day_name'] = df['date'].dt.day_name()

        # Convert time columns to 24-hour format
        df['morning_check_in'] = df['morning_check_in'].apply(
            lambda x: convert_to_24hr_format(str(x)) if pd.notnull(x) else x
        )
        df['last_check_out'] = df['last_check_out'].apply(
            lambda x: convert_to_24hr_format(str(x)) if pd.notnull(x) else x
        )

        # Calculate total hours worked
        df['hours_worked'] = df.apply(calculate_hours_worked, axis=1)

        # Calculate total hours worked by week
        df['week'] = df['date'].dt.isocalendar().week
        weekly_summary = df.groupby(['id', 'week'])['hours_worked'].sum().reset_index()

        # Calculate total hours worked by custom month
        df['custom_month'] = df['date'].apply(assign_custom_month)
        monthly_summary = df.groupby(['id', 'custom_month'])['hours_worked'].sum().reset_index()
        monthly_summary.rename(columns={'custom_month': 'month'}, inplace=True)

        # Create an Excel file in memory using openpyxl
        wb = Workbook()

        # Add the main data to the Excel file
        main_sheet = wb.active
        main_sheet.title = "Main Data"
        for row in dataframe_to_rows(df[['id', 'day_name', 'date', 'morning_check_in', 'last_check_out', 'hours_worked']], index=False, header=True):
            main_sheet.append(row)

        # Add weekly summary to the Excel file
        weekly_sheet = wb.create_sheet("Weekly Summary")
        for row in dataframe_to_rows(weekly_summary, index=False, header=True):
            weekly_sheet.append(row)

        # Add monthly summary to the Excel file
        monthly_sheet = wb.create_sheet("Monthly Summary")
        for row in dataframe_to_rows(monthly_summary, index=False, header=True):
            monthly_sheet.append(row)

        # Create the response object to send the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=uploaded_data.xlsx'

        # Save the workbook to the response
        wb.save(response)

        return response

    # If no file was uploaded or not a POST request, return the upload page
    return render(request, 'upload.html')




